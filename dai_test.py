import openpyxl
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

class ExcelAutomator(object):
    def __init__(self, filename) -> None:
        self.filename = filename
        self.wb = load_workbook(filename)

    def get_sheet(self, sheet_name):
        ws = self.wb[sheet_name]
        return ws
    
    def updated_dai_df(self, ws, df:pd.DataFrame):
        d = df.to_dict(orient='index')

        for cell, data in d.items():
            ws[cell]=data['value']

    def updated_dai(self, ws, metrics:json):

        for metric in d['metrics']:
            cell = metric['this_week_cell']
            ws[cell]=metric['this_week']

            cell = self.get_next_col(cell)
            ws[cell] = metric['last_week']

            cell = self.get_next_col(cell)
            ws[cell] = metric['MTD']

            cell = self.get_next_col(cell)
            ws[cell] = metric['QTD']

            cell = self.get_next_col(cell)
            ws[cell] = metric['target']


    def move_to_last_week(self, ws, df:pd.DataFrame):
        ws.move_range()

    def save(self, filename):
        self.wb.save(filename=filename)

    def get_next_col(self, cell_loc:str)->str:
        #split col from row
        col = ''
        row = ''

        for c in cell_loc:
            if c.isnumeric():
                row = row + c
            elif c.isalpha():
                col = col + c
            else:
                print('Cell location is invalid.')

        next_col = chr(ord(col)+1)

        return next_col + row


if __name__ == "__main__":
    cols = ['date', 'metric', 'value', 'cell']
    data = [['1-1-21', 'Uses Impression Rate (%)', 1.0, 'D33'], 
            ['1-1-21', 'Fill rate, Ad only (%)', 2.3, 'D34'], 
            ['1-1-21', 'Fill rate, Ad+Promo (%)', 3.6, 'D35'], 
            ['1-1-21', 'Completed Impression Rate (%)', 4.9, 'D36']
        ]

    d = {'metrics':[
        {'date' : '1-1-21', 'name' : 'Uses Impression Rate (%)', 'this_week' : 65.89, 'last_week' : 66.51, 'MTD' : 66.35, 'QTD' : 67.06, 'target' : 12, 'this_week_cell' : 'D33'},
        {'date' : '1-1-21', 'name' : 'Fill rate, Ad only (%)', 'this_week' : 1.89, 'last_week' : 11.51, 'MTD' : 111.35, 'QTD' : 1111.06, 'target' : 23, 'this_week_cell' : 'D34'},
        {'date' : '1-1-21', 'name' : 'Fill rate, Ad+Promo (%)', 'this_week' : 2.89, 'last_week' : 22.51, 'MTD' : 222.35, 'QTD' : 2222.06, 'target' : 34, 'this_week_cell' : 'D35'},
        {'date' : '1-1-21', 'name' : 'ompleted Impression Rate (%)', 'this_week' : 3.89, 'last_week' : 33.51, 'MTD' : 333.35, 'QTD' : 3333.06, 'target' : 45, 'this_week_cell' : 'D36'}
    ]}


    metric_json = json.loads(json.dumps(d))

    df = pd.DataFrame(columns=cols, data=data)
    df.index = df['cell']
    df.drop('cell', axis=1)
    filename = 'TeamMetrics.xlsx'

    exa = ExcelAutomator(filename)
    ws = exa.get_sheet('SlingTV')
    exa.updated_dai(ws, metric_json)
    exa.save(filename)

    #print(exa.get_next_col('D33'))
'''
,
            ['1-8-21', 'Uses Impression Rate (%)', 1.1, 'D33'], 
            ['1-8-21', 'Fill rate, Ad only (%)', 2.4, 'D34'], 
            ['1-8-21', 'Fill rate, Ad+Promo (%)', 3.7, 'D35'], 
            ['1-8-21', 'Completed Impression Rate (%)', 5.0, 'D36'],
            ['1-15-21', 'Uses Impression Rate (%)', 0.9, 'D33'], 
            ['1-15-21', 'Fill rate, Ad only (%)', 2.2, 'D34'], 
            ['1-15-21', 'Fill rate, Ad+Promo (%)', 3.5, 'D35'], 
            ['1-15-21', 'Completed Impression Rate (%)', 4.8, 'D36'],
            ['1-22-21', 'Uses Impression Rate (%)', 1.0, 'D33'], 
            ['1-22-21', 'Fill rate, Ad only (%)', 2.3, 'D34'], 
            ['1-22-21', 'Fill rate, Ad+Promo (%)', 3.6, 'D35'], 
            ['1-22-21', 'Completed Impression Rate (%)', 4.9, 'D36'],
            ['1-30-21', 'Uses Impression Rate (%)', 1.3, 'D33'], 
            ['1-30-21', 'Fill rate, Ad only (%)', 2.6, 'D34'], 
            ['1-30-21', 'Fill rate, Ad+Promo (%)', 3.9, 'D35'], 
            ['1-30-21', 'Completed Impression Rate (%)', 5.2, 'D36'],
'''